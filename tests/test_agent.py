"""
Tests for the AI datasheet-filling agent.

Run with:
    python -m pytest tests/ -v
"""

from __future__ import annotations

import json
import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

# Make the repo root importable regardless of how pytest is invoked.
sys.path.insert(0, str(Path(__file__).parent.parent))

from agent.extractor import (
    DataExtractor,
    _DATASHEET_KEYS,
    _LIST_SHEET_COLUMNS,
    _PARAM_ROW_KEYS,
    _REQUIRED_KEYS,
)
from agent.prompts import SYSTEM_PROMPT, build_user_message
from utils.excel_handler import (
    fill_template,
    read_ebom_reference,
    read_template_structure,
)
from utils.pdf_converter import generate_pdf

# ---------------------------------------------------------------------------
# Shared path to the real Form.xlsx template
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).parent.parent
_FORM_TEMPLATE = _REPO_ROOT / "templates" / "Form.xlsx"

# ---------------------------------------------------------------------------
# Shared fixture data matching the new Form.xlsx schema
# ---------------------------------------------------------------------------

_SAMPLE_DATA: dict = {
    "Datasheet": {
        "Author": "Jane Smith",
        "Item Name": "Test Motor 15 kW",
        "HEL": "RMS-01",
        "System Description": "A 15 kW three-phase induction motor for pump drive.",
        "Dimensional Parameters": [
            {"Parameter": "Height",  "Unit": "mm",  "Value": "500",  "Reference": "[1]", "Notes": ""},
            {"Parameter": "Width",   "Unit": "mm",  "Value": "300",  "Reference": "[1]", "Notes": ""},
            {"Parameter": "Length",  "Unit": "mm",  "Value": "800",  "Reference": "[1]", "Notes": ""},
            {"Parameter": "Mass",    "Unit": "kg",  "Value": "150",  "Reference": "[1]", "Notes": ""},
        ],
        "Other Parameters": [
            {"Parameter": "Min Operating Temperature", "Unit": "°C", "Value": "-20",    "Reference": "[1]", "Notes": ""},
            {"Parameter": "Max Operating Temperature", "Unit": "°C", "Value": "80",     "Reference": "[1]", "Notes": ""},
            {"Parameter": "Rated Power",               "Unit": "kW", "Value": "15",     "Reference": "[1]", "Notes": ""},
            {"Parameter": "Supply Voltage",            "Unit": "V",  "Value": "400",    "Reference": "[1]", "Notes": ""},
        ],
        "Manufacturer": "Siemens",
        "Model": "SIMOTICS GP 1LE1",
        "Website": "https://siemens.com/motor",
        "Notes": "1. Values based on standard industrial conditions.",
        "References": "[1] Siemens, 'SIMOTICS GP Motor Datasheet', 2023.",
    },
    "EBOM": [
        {
            "HEL": "RMS-01",
            "Responsible person": "Jane Smith",
            "Task": "Pump Drive",
            "Machine type": "Electric Motor",
            "Specific machine": "SIMOTICS GP 1LE1 15kW",
            "Product website": "https://siemens.com/motor",
            "Product phase": "Off-The-Shelf",
            "Description": "15 kW induction motor for driving the process pump.",
            "Height (mm)": "500",
            "Length (mm)": "800",
            "Width (mm)": "300",
            "Mass (kg)": "150",
            "TRL": "9",
            "SRL": "9",
            "MRL": "10",
        }
    ],
    "SRD": [
        {
            "HEL": "RMS-01",
            "No": "#1",
            "Requirement": "The motor shall deliver 15 kW continuously at rated voltage.",
            "Requirement Type": "Functional requirement",
        },
        {
            "HEL": "RMS-01",
            "No": "#2",
            "Requirement": "The motor shall operate between -20 °C and 80 °C.",
            "Requirement Type": "Constraint",
        },
    ],
    "CDD": [
        {
            "HEL": "RMS-01",
            "No": "#1",
            "Statement": "The motor shall interface with the pump drive shaft.",
        },
        {
            "HEL": "RMS-01",
            "No": "#2",
            "Statement": "The motor shall be powered from the 400 V plant grid.",
        },
    ],
}


# ---------------------------------------------------------------------------
# Tests for agent.prompts
# ---------------------------------------------------------------------------


class TestPrompts(unittest.TestCase):
    def test_system_prompt_is_nonempty(self):
        self.assertGreater(len(SYSTEM_PROMPT), 100)

    def test_system_prompt_contains_sheet_names(self):
        for key in ["Datasheet", "EBOM", "SRD", "CDD"]:
            self.assertIn(key, SYSTEM_PROMPT)

    def test_system_prompt_contains_form_column_names(self):
        for col in ["Responsible person", "Machine type", "Product phase",
                    "Requirement Type", "Statement", "HEL"]:
            self.assertIn(col, SYSTEM_PROMPT)

    def test_build_user_message_contains_machine_name(self):
        msg = build_user_message("Test Motor 5 kW")
        self.assertIn("Test Motor 5 kW", msg)

    def test_build_user_message_with_task_type(self):
        msg = build_user_message("Test Motor", task_type="maintenance")
        self.assertIn("maintenance", msg)

    def test_build_user_message_with_ebom_reference(self):
        ref = [{"HEL": "RMS-01", "Machine type": "Industrial robot"}]
        msg = build_user_message("Test Motor", ebom_reference=ref)
        self.assertIn("Industrial robot", msg)

    def test_build_user_message_no_optional_args(self):
        msg = build_user_message("Motor X")
        self.assertNotIn("Task Type", msg)
        self.assertNotIn("EBOM Reference", msg)


# ---------------------------------------------------------------------------
# Tests for agent.extractor (DataExtractor)
# ---------------------------------------------------------------------------


class TestDataExtractorValidation(unittest.TestCase):
    def _valid_json(self):
        return json.dumps(_SAMPLE_DATA)

    def test_valid_response_parsed_correctly(self):
        result = DataExtractor._parse_and_validate(self._valid_json())
        self.assertEqual(set(result.keys()), _REQUIRED_KEYS)
        self.assertIsInstance(result["Datasheet"], dict)
        self.assertEqual(result["Datasheet"]["Item Name"], "Test Motor 15 kW")

    def test_strips_markdown_fences(self):
        fenced = "```json\n" + self._valid_json() + "\n```"
        result = DataExtractor._parse_and_validate(fenced)
        self.assertIn("Datasheet", result)

    def test_strips_plain_backtick_fences(self):
        fenced = "```\n" + self._valid_json() + "\n```"
        result = DataExtractor._parse_and_validate(fenced)
        self.assertIn("EBOM", result)

    def test_raises_on_invalid_json(self):
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate("not json at all")
        self.assertIn("not valid JSON", str(ctx.exception))

    def test_raises_on_missing_top_level_key(self):
        data = dict(_SAMPLE_DATA)
        del data["CDD"]
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("CDD", str(ctx.exception))

    def test_raises_on_non_dict_response(self):
        with self.assertRaises(ValueError):
            DataExtractor._parse_and_validate("[1, 2, 3]")

    def test_raises_when_datasheet_is_list_not_dict(self):
        data = dict(_SAMPLE_DATA)
        data["Datasheet"] = [{"key": "val"}]
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("Datasheet", str(ctx.exception))

    def test_raises_on_missing_datasheet_key(self):
        data = dict(_SAMPLE_DATA)
        ds = dict(data["Datasheet"])
        del ds["Manufacturer"]
        data = dict(data)
        data["Datasheet"] = ds
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("Manufacturer", str(ctx.exception))

    def test_raises_on_missing_param_row_key(self):
        data = dict(_SAMPLE_DATA)
        ds = dict(data["Datasheet"])
        ds["Dimensional Parameters"] = [{"Parameter": "Height"}]  # missing Unit/Value/Reference/Notes
        data = dict(data)
        data["Datasheet"] = ds
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("Dimensional Parameters", str(ctx.exception))

    def test_raises_on_missing_ebom_column(self):
        data = dict(_SAMPLE_DATA)
        data["EBOM"] = [{"HEL": "RMS-01"}]  # missing many columns
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("EBOM", str(ctx.exception))

    def test_raises_on_missing_srd_column(self):
        data = dict(_SAMPLE_DATA)
        data["SRD"] = [{"HEL": "RMS-01", "No": "#1"}]  # missing Requirement/Requirement Type
        with self.assertRaises(ValueError) as ctx:
            DataExtractor._parse_and_validate(json.dumps(data))
        self.assertIn("SRD", str(ctx.exception))

    def test_raises_when_no_api_key(self):
        # Without a key, cloud providers should raise ValueError/RuntimeError.
        env_without_keys = {
            k: v for k, v in os.environ.items()
            if k not in ("OPENAI_API_KEY", "GOOGLE_API_KEY", "DEEPSEEK_API_KEY",
                         "GROQ_API_KEY", "LLM_PROVIDER")
        }
        with patch.dict(os.environ, env_without_keys, clear=True):
            for provider in ("openai", "google", "deepseek", "groq"):
                extractor = DataExtractor(api_key=None, provider=provider)
                with self.assertRaises((ValueError, RuntimeError),
                                       msg=f"Expected error for provider={provider} with no key"):
                    extractor.extract("Test Motor")

    def test_extract_calls_openai_and_returns_data(self):
        """Mock the OpenAI client so we don't make real API calls."""
        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps(_SAMPLE_DATA)
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = mock_response

        with patch("agent.extractor.OpenAI", return_value=mock_client):
            extractor = DataExtractor(api_key="test-key-123")
            result = extractor.extract("Siemens Motor 15 kW", task_type="maintenance")

        self.assertIn("Datasheet", result)
        self.assertIn("EBOM", result)
        self.assertEqual(result["Datasheet"]["Author"], "Jane Smith")
        self.assertEqual(result["EBOM"][0]["HEL"], "RMS-01")

    def test_extractor_uses_env_model(self):
        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps(_SAMPLE_DATA)
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = mock_response

        with patch.dict(os.environ, {"OPENAI_MODEL": "gpt-4-turbo", "OPENAI_API_KEY": "k"}):
            with patch("agent.extractor.OpenAI", return_value=mock_client):
                extractor = DataExtractor()
                extractor.extract("Motor")
                call_kwargs = mock_client.chat.completions.create.call_args[1]
                self.assertEqual(call_kwargs["model"], "gpt-4-turbo")


# ---------------------------------------------------------------------------
# Tests for utils.excel_handler — filling Form.xlsx
# ---------------------------------------------------------------------------


class TestExcelHandler(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.mkdtemp()
        self._output_path = Path(self._tmpdir) / "filled.xlsx"

    def _require_form_template(self):
        if not _FORM_TEMPLATE.exists():
            self.skipTest("templates/Form.xlsx not present")
        return _FORM_TEMPLATE

    def test_read_template_structure_returns_ebom_headers(self):
        template = self._require_form_template()
        structure = read_template_structure(template)
        self.assertIn("EBOM", structure)
        self.assertIn("HEL", structure["EBOM"])

    def test_read_template_structure_returns_all_sheets(self):
        template = self._require_form_template()
        structure = read_template_structure(template)
        for sheet in ("EBOM", "Datasheet", "SRD", "CDD"):
            self.assertIn(sheet, structure)

    def test_fill_template_creates_output_file(self):
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        self.assertTrue(self._output_path.exists())

    def test_fill_template_datasheet_author(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["Datasheet"]
        # B3 should contain the author line
        self.assertIn("Jane Smith", str(ws["B3"].value or ""))

    def test_fill_template_datasheet_item_name(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["Datasheet"]
        self.assertIn("Test Motor", str(ws["E5"].value or ""))

    def test_fill_template_datasheet_hel(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["Datasheet"]
        self.assertEqual(ws["E6"].value, "RMS-01")

    def test_fill_template_datasheet_dim_param_row(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["Datasheet"]
        # Row 17 = first dimensional parameter (Height)
        self.assertEqual(str(ws.cell(row=17, column=2).value), "Height")
        self.assertEqual(str(ws.cell(row=17, column=6).value), "500")

    def test_fill_template_datasheet_manufacturer(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["Datasheet"]
        self.assertEqual(ws["E33"].value, "Siemens")

    def test_fill_template_ebom_row(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["EBOM"]
        # Row 4 = first new EBOM machine (RMS-1 slot)
        self.assertEqual(ws.cell(row=4, column=2).value, "Jane Smith")   # Responsible person
        self.assertEqual(ws.cell(row=4, column=7).value, "Off-The-Shelf")  # Product phase

    def test_fill_template_srd_rows(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["SRD"]
        self.assertEqual(ws.cell(row=3, column=1).value, "RMS-01")   # HEL
        self.assertEqual(ws.cell(row=3, column=2).value, "#1")         # No

    def test_fill_template_cdd_rows(self):
        import openpyxl
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        wb = openpyxl.load_workbook(self._output_path, data_only=True)
        ws = wb["CDD"]
        self.assertEqual(ws.cell(row=3, column=3).value,
                         "The motor shall interface with the pump drive shaft.")

    def test_fill_template_creates_output_dir(self):
        template = self._require_form_template()
        deep_output = Path(self._tmpdir) / "nested" / "deep" / "filled.xlsx"
        fill_template(_SAMPLE_DATA, template, deep_output)
        self.assertTrue(deep_output.exists())

    def test_read_ebom_reference_from_filled_template(self):
        template = self._require_form_template()
        fill_template(_SAMPLE_DATA, template, self._output_path)
        rows = read_ebom_reference(self._output_path, sheet_name="EBOM")
        self.assertGreater(len(rows), 0)

    def test_read_ebom_reference_missing_sheet(self):
        template = self._require_form_template()
        rows = read_ebom_reference(template, sheet_name="NONEXISTENT")
        self.assertEqual(rows, [])

    def test_template_file_not_modified(self):
        """Ensure fill_template never modifies the original template."""
        template = self._require_form_template()
        mtime_before = template.stat().st_mtime
        fill_template(_SAMPLE_DATA, template, self._output_path)
        self.assertEqual(template.stat().st_mtime, mtime_before)


# ---------------------------------------------------------------------------
# Tests for utils.pdf_converter
# ---------------------------------------------------------------------------


class TestPdfConverter(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.mkdtemp()
        self._pdf_path = Path(self._tmpdir) / "output.pdf"

    def test_generate_pdf_creates_file(self):
        generate_pdf(_SAMPLE_DATA, self._pdf_path, machine_name="Test Motor")
        self.assertTrue(self._pdf_path.exists())
        self.assertGreater(self._pdf_path.stat().st_size, 1000)

    def test_generate_pdf_creates_parent_dirs(self):
        deep_pdf = Path(self._tmpdir) / "a" / "b" / "report.pdf"
        generate_pdf(_SAMPLE_DATA, deep_pdf)
        self.assertTrue(deep_pdf.exists())

    def test_generate_pdf_with_empty_tabular_sections(self):
        sparse_data = {
            "Datasheet": _SAMPLE_DATA["Datasheet"],
            "EBOM": [],
            "SRD": [],
            "CDD": [],
        }
        generate_pdf(sparse_data, self._pdf_path)
        self.assertTrue(self._pdf_path.exists())

    def test_generate_pdf_returns_path(self):
        result = generate_pdf(_SAMPLE_DATA, self._pdf_path)
        self.assertEqual(result, self._pdf_path)


# ---------------------------------------------------------------------------
# Integration-style test for main.run (mocked LLM)
# ---------------------------------------------------------------------------


class TestMainRun(unittest.TestCase):
    def setUp(self):
        self._tmpdir = tempfile.mkdtemp()

    def _require_form_template(self):
        if not _FORM_TEMPLATE.exists():
            self.skipTest("templates/Form.xlsx not present")
        return _FORM_TEMPLATE

    def test_run_full_pipeline(self):
        """Integration test: mock LLM, run full pipeline, verify outputs."""
        from main import run

        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps(_SAMPLE_DATA)
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = mock_response

        output_dir = Path(self._tmpdir) / "output"
        template_path = self._require_form_template()

        with patch("agent.extractor.OpenAI", return_value=mock_client):
            with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
                result = run([
                    "--machine", "Test Motor 5 kW",
                    "--task", "maintenance",
                    "--template", str(template_path),
                    "--output", str(output_dir),
                ])

        self.assertIn("Datasheet", result)
        xlsx_files = list(output_dir.glob("*.xlsx"))
        self.assertEqual(len(xlsx_files), 1)
        pdf_files = list(output_dir.glob("*.pdf"))
        self.assertEqual(len(pdf_files), 1)

    def test_run_no_pdf_flag(self):
        """--no-pdf should skip PDF generation."""
        from main import run

        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps(_SAMPLE_DATA)
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = mock_response

        output_dir = Path(self._tmpdir) / "output_nopdf"
        template_path = self._require_form_template()

        with patch("agent.extractor.OpenAI", return_value=mock_client):
            with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
                run([
                    "--machine", "Test Motor",
                    "--template", str(template_path),
                    "--output", str(output_dir),
                    "--no-pdf",
                ])

        pdf_files = list(output_dir.glob("*.pdf"))
        self.assertEqual(len(pdf_files), 0)

    def test_run_saves_json_when_requested(self):
        from main import run

        mock_response = MagicMock()
        mock_response.choices[0].message.content = json.dumps(_SAMPLE_DATA)
        mock_client = MagicMock()
        mock_client.chat.completions.create.return_value = mock_response

        output_dir = Path(self._tmpdir) / "output_json"
        template_path = self._require_form_template()
        json_out = Path(self._tmpdir) / "out.json"

        with patch("agent.extractor.OpenAI", return_value=mock_client):
            with patch.dict(os.environ, {"OPENAI_API_KEY": "test-key"}):
                run([
                    "--machine", "Test Motor",
                    "--template", str(template_path),
                    "--output", str(output_dir),
                    "--no-pdf",
                    "--json-out", str(json_out),
                ])

        self.assertTrue(json_out.exists())
        loaded = json.loads(json_out.read_text())
        self.assertIn("Datasheet", loaded)


if __name__ == "__main__":
    unittest.main()
