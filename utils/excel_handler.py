"""
Excel handler utilities — fill the Form.xlsx template with agent data, and
read EBOM reference rows from existing XLSX files.

The Form.xlsx template has five sheets:
  • EBOM          — tabular, rows 3-21 (row 3 = demo/example RMS-0)
  • Datasheet     — structured form with merged cells and parameter tables
  • SRD           — tabular, rows 3+
  • CDD           — tabular, rows 3+
  • Drop Down List — lookup values for Product phase (do not modify)

Public functions
----------------
fill_template(data, template_path, output_path)
    Copy the Form.xlsx template, clear editable cells, and fill with data.

read_template_structure(path)
    Return the header columns for each sheet (informational).

read_ebom_reference(path, sheet_name)
    Return rows from an EBOM sheet in an existing XLSX as a list of dicts.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment

# ---------------------------------------------------------------------------
# Form.xlsx layout constants
# ---------------------------------------------------------------------------

# EBOM: columns to fill (1-indexed), data rows start at 4 (row 3 = example).
_EBOM_COL_MAP: dict[str, int] = {
    "Responsible person": 2,   # B
    "Task":               3,   # C
    "Machine type":       4,   # D
    "Specific machine":   5,   # E
    "Product website":    6,   # F
    "Product phase":      7,   # G
    "Description":        8,   # H
    "Height (mm)":        9,   # I
    "Length (mm)":       10,   # J
    "Width (mm)":        11,   # K
    "Mass (kg)":         12,   # L
    "TRL":               13,   # M
    "SRL":               14,   # N
    "MRL":               15,   # O
}
_EBOM_DATA_START_ROW = 4   # row 3 is the pre-filled RMS-0 example

# Datasheet: exact cell addresses for scalar fields.
# Write to the *first* cell of each merged range (openpyxl rule).
_DS_AUTHOR_CELL      = "B3"
_DS_ITEM_NAME_CELL   = "E5"
_DS_HEL_CELL         = "E6"
_DS_DESC_CELL        = "B9"

# Dimensional parameter table rows (rows 17-20), sub-section "a) Dimensional…"
_DS_DIM_PARAM_ROWS   = [17, 18, 19, 20]

# "b) Other Information" sub-section header is at row 21.
# Other parameter rows follow from row 22.
_DS_OTHER_PARAM_START = 22
_DS_OTHER_PARAM_MAX   = 9   # rows 22-30 (9 rows available in the blank form)

# Columns inside the parameter table (first cell of each merge):
_DS_PARAM_COL  = 2   # B — parameter name
_DS_UNIT_COL   = 4   # D — unit
_DS_VALUE_COL  = 6   # F — value
_DS_REF_COL    = 8   # H — reference
_DS_NOTES_COL  = 10  # J — notes

_DS_MANUFACTURER_CELL = "E33"
_DS_MODEL_CELL        = "E34"
_DS_WEBSITE_CELL      = "E35"
_DS_NOTES_CELL        = "B38"
_DS_REFERENCES_CELL   = "B41"

# SRD / CDD: tabular, data rows start at row 3.
_SRD_COL_MAP: dict[str, int] = {
    "HEL":              1,   # A
    "No":               2,   # B
    "Requirement":      3,   # C
    "Requirement Type": 4,   # D
}

_CDD_COL_MAP: dict[str, int] = {
    "HEL":       1,   # A
    "No":        2,   # B
    "Statement": 3,   # C
}

_TABULAR_DATA_START_ROW = 3

# All supported form keys
_ALL_FORMS = ["Datasheet", "EBOM", "SRD", "CDD"]

# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def fill_template(
    data: dict[str, Any],
    template_path: str | Path,
    output_path: str | Path,
    forms: Optional[list[str]] = None,
    image_path: Optional[Path] = None,
) -> Path:
    """Fill *template_path* (Form.xlsx) with *data* and save to *output_path*.

    Parameters
    ----------
    data:
        Dict produced by :class:`agent.extractor.DataExtractor`. Keys:
        ``Datasheet`` (dict), ``EBOM``, ``SRD``, ``CDD`` (lists of dicts).
    template_path:
        Path to the blank Form.xlsx template.
    output_path:
        Destination path for the filled XLSX.
    forms:
        List of sheet names to fill. Defaults to all four sheets.
        Valid values: ``"Datasheet"``, ``"EBOM"``, ``"SRD"``, ``"CDD"``.
    image_path:
        Optional path to a machine image file to embed in the Datasheet sheet.

    Returns
    -------
    Path
        Resolved path of the written file.
    """
    if forms is None:
        forms = _ALL_FORMS

    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Work on a fresh copy of the template so the original stays untouched.
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)

    if "Datasheet" in forms:
        _fill_datasheet(wb, data.get("Datasheet", {}))
        if image_path is not None:
            _embed_image(wb, image_path)
    if "EBOM" in forms:
        _fill_ebom(wb, data.get("EBOM", []))
    if "SRD" in forms:
        _fill_srd(wb, data.get("SRD", []))
    if "CDD" in forms:
        _fill_cdd(wb, data.get("CDD", []))

    wb.save(output_path)
    return output_path


def read_template_structure(path: str | Path) -> dict[str, list[str]]:
    """Return the header columns for every sheet in the template XLSX.

    Parameters
    ----------
    path:
        Path to the template XLSX.

    Returns
    -------
    dict
        Mapping of sheet name → list of column header strings (first row).
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    structure: dict[str, list[str]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(cell) for cell in row if cell is not None]
        structure[sheet_name] = headers
    wb.close()
    return structure


def read_ebom_reference(
    path: str | Path,
    sheet_name: str = "EBOM",
) -> list[dict[str, Any]]:
    """Read EBOM reference data from an XLSX file.

    Parameters
    ----------
    path:
        Path to the XLSX file containing EBOM reference data.
    sheet_name:
        Name of the sheet to read (default: "EBOM").

    Returns
    -------
    list of dict
        One dict per data row, keyed by column header.
    """
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        wb.close()
        return []

    records: list[dict[str, Any]] = []
    for row in rows_iter:
        record = {headers[i]: (row[i] if i < len(row) else None)
                  for i in range(len(headers))}
        records.append(record)

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Private — sheet-filling helpers
# ---------------------------------------------------------------------------

_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="center")


def _set(ws, cell_ref: str, value: Any) -> None:
    """Write *value* to *cell_ref*, preserving existing style."""
    cell = ws[cell_ref]
    cell.value = str(value) if value is not None else ""
    cell.alignment = _WRAP_TOP


def _set_rc(ws, row: int, col: int, value: Any) -> None:
    """Write *value* to row/col (1-indexed), preserving existing style."""
    cell = ws.cell(row=row, column=col)
    cell.value = str(value) if value is not None else ""
    cell.alignment = _WRAP_TOP


# ---- Image embedding -------------------------------------------------------

def _embed_image(wb: openpyxl.Workbook, image_path: Path) -> None:
    """Embed *image_path* into the Datasheet sheet at cell H3."""
    if "Datasheet" not in wb.sheetnames:
        return
    try:
        from openpyxl.drawing.image import Image as XLImage
        ws = wb["Datasheet"]
        img = XLImage(str(image_path))
        # Scale to a reasonable thumbnail (approx 160×120 px)
        img.width = 160
        img.height = 120
        ws.add_image(img, "H3")
    except Exception as exc:  # pragma: no cover
        import logging
        logging.getLogger(__name__).warning("Could not embed image in XLSX: %s", exc)


# ---- Datasheet -------------------------------------------------------------

def _fill_datasheet(wb: openpyxl.Workbook, ds: dict[str, Any]) -> None:
    """Fill the Datasheet sheet from *ds*."""
    if "Datasheet" not in wb.sheetnames:
        return
    ws = wb["Datasheet"]

    _set(ws, _DS_AUTHOR_CELL, f"Author: {ds.get('Author', '')}")
    _set(ws, _DS_ITEM_NAME_CELL, ds.get("Item Name", ""))
    _set(ws, _DS_HEL_CELL, ds.get("HEL", ""))
    _set(ws, _DS_DESC_CELL, ds.get("System Description", ""))

    # Dimensional parameters (rows 17-20)
    dim_params = ds.get("Dimensional Parameters", [])
    for i, row_num in enumerate(_DS_DIM_PARAM_ROWS):
        if i < len(dim_params):
            p = dim_params[i]
            _set_rc(ws, row_num, _DS_PARAM_COL, p.get("Parameter", ""))
            _set_rc(ws, row_num, _DS_UNIT_COL,  p.get("Unit", ""))
            _set_rc(ws, row_num, _DS_VALUE_COL, p.get("Value", ""))
            _set_rc(ws, row_num, _DS_REF_COL,   p.get("Reference", ""))
            _set_rc(ws, row_num, _DS_NOTES_COL, p.get("Notes", ""))
        else:
            # Clear the row if no data provided
            for col in (_DS_PARAM_COL, _DS_UNIT_COL, _DS_VALUE_COL,
                        _DS_REF_COL, _DS_NOTES_COL):
                _set_rc(ws, row_num, col, "")

    # Other parameters (rows 22 .. 22+max-1)
    other_params = ds.get("Other Parameters", [])
    for i in range(_DS_OTHER_PARAM_MAX):
        row_num = _DS_OTHER_PARAM_START + i
        if i < len(other_params):
            p = other_params[i]
            _set_rc(ws, row_num, _DS_PARAM_COL, p.get("Parameter", ""))
            _set_rc(ws, row_num, _DS_UNIT_COL,  p.get("Unit", ""))
            _set_rc(ws, row_num, _DS_VALUE_COL, p.get("Value", ""))
            _set_rc(ws, row_num, _DS_REF_COL,   p.get("Reference", ""))
            _set_rc(ws, row_num, _DS_NOTES_COL, p.get("Notes", ""))
        else:
            for col in (_DS_PARAM_COL, _DS_UNIT_COL, _DS_VALUE_COL,
                        _DS_REF_COL, _DS_NOTES_COL):
                _set_rc(ws, row_num, col, "")

    _set(ws, _DS_MANUFACTURER_CELL, ds.get("Manufacturer", ""))
    _set(ws, _DS_MODEL_CELL,        ds.get("Model", ""))
    _set(ws, _DS_WEBSITE_CELL,      ds.get("Website", ""))
    _set(ws, _DS_NOTES_CELL,        ds.get("Notes", ""))
    _set(ws, _DS_REFERENCES_CELL,   ds.get("References", ""))


# ---- EBOM ------------------------------------------------------------------

def _fill_ebom(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    """Fill EBOM data rows starting at row 4 (row 3 = example RMS-0)."""
    if "EBOM" not in wb.sheetnames:
        return
    ws = wb["EBOM"]

    for i, row_data in enumerate(rows):
        row_num = _EBOM_DATA_START_ROW + i
        # Column A (HEL) is pre-filled in the template (RMS-1, RMS-2, …).
        # Overwrite it with the HEL from the data if provided.
        hel_val = row_data.get("HEL", "")
        if hel_val:
            _set_rc(ws, row_num, 1, hel_val)
        for field, col in _EBOM_COL_MAP.items():
            _set_rc(ws, row_num, col, row_data.get(field, ""))


# ---- SRD -------------------------------------------------------------------

def _fill_srd(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    """Fill SRD data rows starting at row 3."""
    if "SRD" not in wb.sheetnames:
        return
    ws = wb["SRD"]

    for i, row_data in enumerate(rows):
        row_num = _TABULAR_DATA_START_ROW + i
        for field, col in _SRD_COL_MAP.items():
            _set_rc(ws, row_num, col, row_data.get(field, ""))


# ---- CDD -------------------------------------------------------------------

def _fill_cdd(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    """Fill CDD data rows starting at row 3."""
    if "CDD" not in wb.sheetnames:
        return
    ws = wb["CDD"]

    for i, row_data in enumerate(rows):
        row_num = _TABULAR_DATA_START_ROW + i
        for field, col in _CDD_COL_MAP.items():
            _set_rc(ws, row_num, col, row_data.get(field, ""))
